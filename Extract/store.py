from openpyxl import Workbook
from openpyxl import load_workbook
import pymysql
import xlrd
import xlwt
import os
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from xlwt import Workbook
import xlsxwriter
ft = Font("FFBB00")
Alist=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ','DA','DB','DC','DD','DE','DF','DG','DH','DI','DJ','DK','DL','DM','DN','DO','DP','DQ','DR','DS','DT','DU','DV','DW','DX','DY','DZ']
# def get_data():
	# conn = pymysql.connect(host='127.0.0.1',port=3306,user='root',passwd='123456',db='%s'%(dbName))
	# cursor = conn.cursor()
	# sql = "SELECT url FROM htm_title WHERE title='title'"
	# cursor.execute(sql) 
	# conn.commit() 
	# data = cursor.fetchall()
	# return data
def table_name(host,port,tableName,dbName):
	conn = pymysql.connect(host=host,port=port,user='root',passwd='123456',db='%s'%(dbName))
	cursor = conn.cursor()	
	#sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = 'hhc' and table_schema = 'work770';"
	sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s' and table_schema = '%s'"%(tableName,dbName)
	cursor.execute(sql) 
	conn.commit() 
	data = cursor.fetchall()
	dataList=['id','JPN','ENU']
	for i in data:
		if i[0] not in ['id','JPN','ENU','tableName']:
			dataList.append(i[0])
	dataList.append('tableName')
	return dataList
# def show_tables():
	# conn = pymysql.connect(host='127.0.0.1',port=3306,user='root',passwd='123456',db='%s'%(dbName))
	# cursor = conn.cursor()
	# sql = "show tables"
	# cursor.execute(sql) 
	# conn.commit() 
	# data = cursor.fetchall()
	#print(data)
	#return data
def createXls():
	book = Workbook(encoding='utf-8')
	for tableName in tableNameList:	
		sheet1 = book.add_sheet('%s'%(tableName))
	book.save('%s.xlsx'%(dbName))
def write_excel(host,port,colName,tableName,excelName,dbName,dir_choose):
	#colList
	#rowNum="select count('url') from %s"%(tableName)
	#tableName
	conn = pymysql.connect(host=host,port=port,user='root',passwd='123456',db='%s'%(dbName))
	cursor = conn.cursor()
	storeName=dbName+'+'+tableName
	wb = load_workbook(r"%s/%s.xlsx"%(dir_choose,storeName))

	ws = wb.get_sheet_by_name('%s'%(storeName))
	#print(colName,tableName)
	sql = "SELECT %s FROM %s"%(colName,tableName)
	print(sql)
	cursor.execute(sql) 
	conn.commit() 
	ws['%s1'%(Alist[excelName])] =colName
	coldata= cursor.fetchall()
	#print(coldata)
	for colnum in range(len(coldata)):
		excelcol=Alist[excelName]
		ws['%s%s'%(excelcol,str(colnum+2))] =coldata[colnum][0]
	# Row=1
	# for k,v in enumerate(tunple):
		# ws['A%s'%(str(k+1))] =tunple[k][0]
		
	wb.save(r"%s\%s.xlsx"%(dir_choose,storeName))
def create_xls(dbName,tableNameList):
	workbook = xlsxwriter.Workbook(r"%s.xlsx"%(dbName))
	for tableName in tableNameList:	
		worksheet = workbook.add_worksheet(tableName)
	workbook.close()

def store_xls(host,port,dbName,tableNameList,dir_choose):    
# tunple=get_data()
# print(tunple)
# write()
	#tableNameList=['hhc','hhk','dialog_xml','stringdefinition','htm']
	#dbName='work770'
	#create_xls(dbName,tableNameList)
	for tableName in tableNameList:
		tableName=dbName+'+'+tableName
		workbook = xlsxwriter.Workbook(r"%s/%s.xlsx"%(dir_choose,tableName))
		worksheet = workbook.add_worksheet(tableName)
		workbook.close()
	for tableName in tableNameList:
		
		#tableName='hhc'
		#dbName='work770'
		#wb = load_workbook(r"%s.xlsx"%(dbName))#打开excel
		#ws = wb.get_sheet_by_name('%s'%(tableName))#打开sheet‘IDC流量’
		colList=table_name(host,port,tableName,dbName)      #表的列名
		print(colList)
		for i in range(len(colList)):
			write_excel(host,port,colList[i],tableName,i,dbName,dir_choose)
		print(tableName)



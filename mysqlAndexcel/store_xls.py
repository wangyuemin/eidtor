from openpyxl import Workbook
from openpyxl import load_workbook
import pymysql
import xlrd
import xlwt
import os
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from xlwt import Workbook
import xlsxwriter
ft = Font("FFBB00")
Alist=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']
def get_data():
	conn = pymysql.connect(host='127.0.0.1',port=3306,user='root',passwd='123456',db='work')
	cursor = conn.cursor()
	sql = "SELECT url FROM htm_title WHERE title='title'"
	cursor.execute(sql) 
	conn.commit() 
	data = cursor.fetchall()
	return data
def table_name(tableName,dbName):
	conn = pymysql.connect(host='127.0.0.1',port=3306,user='root',passwd='123456',db='work')
	cursor = conn.cursor()	
	#sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = 'hhc' and table_schema = 'work';"
	sql="select COLUMN_NAME from information_schema.COLUMNS where table_name = '%s' and table_schema = '%s'"%(tableName,dbName)
	cursor.execute(sql) 
	conn.commit() 
	data = cursor.fetchall()
	dataList=[]
	for i in data:
		dataList.append(i[0])
	return dataList
def show_tables():
	conn = pymysql.connect(host='127.0.0.1',port=3306,user='root',passwd='123456',db='work')
	cursor = conn.cursor()
	sql = "show tables"
	cursor.execute(sql) 
	conn.commit() 
	data = cursor.fetchall()
	#print(data)
	#return data
def createXls():
	book = Workbook(encoding='utf-8')
	for tableName in tableNameList:	
		sheet1 = book.add_sheet('%s'%(tableName))
	book.save('%s.xlsx'%(dbName))
def write_excel(colName,tableName,excelName,dbName):
	#colList
	#rowNum="select count('url') from %s"%(tableName)
	#tableName
	conn = pymysql.connect(host='127.0.0.1',port=3306,user='root',passwd='123456',db='work')
	cursor = conn.cursor()
	wb = load_workbook(r"%s.xlsx"%(tableName))
	ws = wb.get_sheet_by_name('%s'%(tableName))
	#print(colName,tableName)
	sql = "SELECT %s FROM %s"%(colName,tableName)
	print(sql)
	cursor.execute(sql) 
	conn.commit() 
	ws['%s1'%(Alist[excelName])] =colName
	coldata= cursor.fetchall()
	#print(coldata)
	for colnum in range(len(coldata)):
		excelcol=Alist[excelName]
		ws['%s%s'%(excelcol,str(colnum+2))] =coldata[colnum][0]
	# Row=1
	# for k,v in enumerate(tunple):
		# ws['A%s'%(str(k+1))] =tunple[k][0]
	wb.save(r".\%s.xlsx"%(tableName))
def create_xls(dbName,tableNameList):
	workbook = xlsxwriter.Workbook(r"%s.xlsx"%(dbName))
	for tableName in tableNameList:	
		worksheet = workbook.add_worksheet(tableName)
	workbook.close()
#dbName1='work'
tableNameList1=['hhc','hhk','htm','dialog_xml','dialog_xml_font','stringdefinition']
#def store_xls(dbName,tableNameList):    
def store_xls(dbName,tableNameList=tableNameList1): 
	for tableName in tableNameList:
		workbook = xlsxwriter.Workbook(r"%s.xlsx"%(tableName))
		worksheet = workbook.add_worksheet(tableName)
		workbook.close()
	for tableName in tableNameList:
		#tableName='hhc'
		#dbName='work'
		#wb = load_workbook(r"%s.xlsx"%(dbName))#打开excel
		#ws = wb.get_sheet_by_name('%s'%(tableName))#打开sheet‘IDC流量’
		colList=table_name(tableName,dbName)      #表的列名
		#print(colList)
		for i in range(len(colList)):
			write_excel(colList[i],tableName,i,dbName)
		print(tableName)
#store_xls('work',['hhc','hhk','htm','dialog_xml_ids','dialog_xml_new','dialog_xml_change'])

